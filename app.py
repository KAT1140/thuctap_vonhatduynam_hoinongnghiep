from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import MySQLdb
import MySQLdb.cursors
from config import config
import os
import glob
from openpyxl import load_workbook, Workbook
import io
from datetime import datetime
from datetime import datetime

app = Flask(__name__)

# Normalize environment name and always fall back to development config.
flask_env = (os.environ.get('FLASK_ENV') or 'development').strip().lower()
if flask_env not in config:
    flask_env = 'development'
app.config.from_object(config[flask_env])

print(f"DEBUG: Environment={flask_env}")
print(f"DEBUG: MYSQL_HOST={app.config.get('MYSQL_HOST')}")
print(f"DEBUG: MYSQL_PORT={app.config.get('MYSQL_PORT')}")
print(f"DEBUG: MYSQL_USER={app.config.get('MYSQL_USER')}")
print(f"DEBUG: MYSQL_DB={app.config.get('MYSQL_DB')}")

# Add Jinja2 globals for pagination
app.jinja_env.globals.update(max=max, min=min, range=range)


# Upload config
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
MAX_CONTENT_LENGTH = 16 * 1024 * 1024  # 16MB max file size

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Database connection
def get_db_connection():
    """Tạo kết nối MySQL"""
    try:
        db_host = app.config.get('MYSQL_HOST') or 'localhost'
        db_port = app.config.get('MYSQL_PORT') or 3306
        db_user = app.config.get('MYSQL_USER') or 'root'
        db_password = app.config.get('MYSQL_PASSWORD')
        if db_password is None:
            db_password = ''
        db_name = app.config.get('MYSQL_DB') or 'hoi_nong_dan'

        print(f"DEBUG: Connecting to {db_host}:{db_port} with user {db_user}")

        conn = MySQLdb.connect(
            host=db_host,
            port=db_port,
            user=db_user,
            passwd=db_password,
            db=db_name,
            charset='utf8mb4',
            cursorclass=MySQLdb.cursors.DictCursor
        )
        return conn
    except MySQLdb.Error as e:
        print(f"Error connecting to database: {e}")
        return None

# Decorator kiểm tra đăng nhập
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Vui lòng đăng nhập', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'error': 'Vui lòng đăng nhập'}), 401
            flash('Vui lòng đăng nhập', 'warning')
            return redirect(url_for('login'))
        
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("SELECT role FROM users WHERE id = %s", (session['user_id'],))
            user = cursor.fetchone()
            cursor.close()
            conn.close()
            
            if not user or user['role'] != 'admin':
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                    return jsonify({'error': 'Bạn không có quyền truy cập'}), 403
                flash('Bạn không có quyền truy cập', 'danger')
                return redirect(url_for('dashboard'))
        
        return f(*args, **kwargs)
    return decorated_function


def roles_required(*allowed_roles):
    """Decorator để check role của user"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('Vui lòng đăng nhập', 'warning')
                return redirect(url_for('login'))
            
            user_role = session.get('role')
            if user_role not in allowed_roles:
                flash('Bạn không có quyền truy cập', 'danger')
                return redirect(url_for('dashboard'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator


def validate_org_parent(org_type, parent):
    """Kiểm tra quan hệ cha-con đúng theo thứ bậc: xã -> chi hội -> tổ hội
    - xã: không có parent (cấp cao nhất hành chính)
    - chi hội: xã làm parent
    - tổ hội: chi hội làm parent"""
    if org_type == 'xa':
        return parent is None
    if org_type == 'chi_hoi':
        return parent is not None and parent['org_type'] == 'xa'
    if org_type == 'to_hoi':
        return parent is not None and parent['org_type'] == 'chi_hoi'
    return False


# HELPER FUNCTIONS - Quản lý hội viên thuộc nhiều tổ hội
def get_member_organizations(member_id, conn=None):
    """Lấy danh sách tất cả tổ chức mà hội viên thuộc về (many-to-many)"""
    if not conn:
        conn = get_db_connection()
        close_conn = True
    else:
        close_conn = False
    
    orgs = []
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT mo.id, mo.organization_id, mo.join_date, mo.role_in_org,
                   o.name, o.org_type
            FROM member_organizations mo
            LEFT JOIN organizations o ON mo.organization_id = o.id
            WHERE mo.member_id = %s
            ORDER BY o.org_type, o.name
        """, (member_id,))
        orgs = cursor.fetchall()
        cursor.close()
        if close_conn:
            conn.close()
    
    return orgs


def add_member_to_organization(member_id, org_id, join_date=None, role=None, conn=None):
    """Thêm hội viên vào 1 tổ chức (không đơn (many-to-many table)"""
    if not conn:
        conn = get_db_connection()
        close_conn = True
    else:
        close_conn = False
    
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO member_organizations (member_id, organization_id, join_date, role_in_org)
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE role_in_org = %s
            """, (member_id, org_id, join_date, role, role))
            conn.commit()
            cursor.close()
            if close_conn:
                conn.close()
            return True
        except Exception as e:
            print(f"Error adding member to organization: {e}")
            if close_conn:
                conn.close()
            return False
    return False


def remove_member_from_organization(member_id, org_id, conn=None):
    """Xóa hội viên khỏi 1 tổ chức (many-to-many table)"""
    if not conn:
        conn = get_db_connection()
        close_conn = True
    else:
        close_conn = False
    
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("""
                DELETE FROM member_organizations
                WHERE member_id = %s AND organization_id = %s
            """, (member_id, org_id))
            conn.commit()
            cursor.close()
            if close_conn:
                conn.close()
            return True
        except Exception as e:
            print(f"Error removing member from organization: {e}")
            if close_conn:
                conn.close()
            return False
    return False


# Routes
@app.route('/')
def index():
    """Trang chủ"""
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Đăng nhập"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if not username or not password:
            flash('Vui lòng nhập tài khoản và mật khẩu', 'warning')
            return redirect(url_for('login'))
        
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM users WHERE username = %s AND is_active = TRUE", (username,))
            user = cursor.fetchone()
            cursor.close()
            conn.close()
            
            if user and check_password_hash(user['password'], password):
                session['user_id'] = user['id']
                session['username'] = user['username']
                session['role'] = user['role']
                session['full_name'] = user['full_name']
                session.permanent = True
                flash(f'Chào mừng {user["full_name"]}!', 'success')
                return redirect(url_for('dashboard'))
            else:
                flash('Tài khoản hoặc mật khẩu không đúng', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """Đăng xuất"""
    session.clear()
    flash('Bạn đã đăng xuất', 'info')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    """Trang dashboard"""
    conn = get_db_connection()
    stats = {}
    
    if conn:
        cursor = conn.cursor()
        
        # Lấy thống kê
        cursor.execute("SELECT COUNT(*) as count FROM members WHERE status = 'active'")
        stats['active_members'] = cursor.fetchone()['count']
        
        cursor.execute("SELECT COUNT(*) as count FROM organizations")
        stats['organizations'] = cursor.fetchone()['count']
        
        cursor.execute("SELECT COUNT(*) as count FROM members")
        stats['reports'] = cursor.fetchone()['count']
        
        cursor.execute("SELECT COUNT(*) as count FROM users")
        stats['users'] = cursor.fetchone()['count']
        
        cursor.close()
        conn.close()
    
    return render_template('dashboard.html', stats=stats)

# Members management
@app.route('/members')
@login_required
@roles_required('admin', 'to_hoi', 'hoi_vien')
def members_list():
    """Danh sách hội viên"""
    conn = get_db_connection()
    members = []
    member_orgs = []
    open_add_member = (request.args.get('open_add') or '').strip() == '1'

    search = (request.args.get('q') or '').strip()
    member_type = (request.args.get('member_type') or '').strip()
    education_level = (request.args.get('education_level') or '').strip()
    status = (request.args.get('status') or '').strip()
    organization_id = (request.args.get('organization_id') or '').strip()  # Chi Hội
    to_hoi_id = (request.args.get('to_hoi_id') or '').strip()  # Tổ Hội
    gender = (request.args.get('gender') or '').strip()
    
    # Initialize pagination variables
    page = int(request.args.get('page', 1))
    per_page = 50
    total_members = 0
    total_pages = 0
    
    # Prepare organization filter (handle Xã and Chi Hội which should include all child organizations)
    org_filter_ids = []
    if to_hoi_id:
        # Direct filter by Tổ Hội
        try:
            org_filter_ids = [int(to_hoi_id)]
        except:
            pass
    elif organization_id:
        # Filter by Chi Hội -> get all to_hoi under it
        try:
            org_id = int(organization_id)
            if conn:
                cursor = conn.cursor()
                cursor.execute("SELECT org_type FROM organizations WHERE id = %s", (org_id,))
                org_result = cursor.fetchone()
                
                if org_result and org_result['org_type'] == 'chi_hoi':
                    # Get all to_hoi under this chi_hoi
                    cursor.execute("""
                        SELECT id FROM organizations 
                        WHERE parent_id = %s AND org_type = 'to_hoi'
                    """, (org_id,))
                    org_filter_ids = [row['id'] for row in cursor.fetchall()]
                cursor.close()
        except:
            pass
    
    if conn:
        cursor = conn.cursor()
        query = """
            SELECT m.*,
                   t.name as chi_hoi_name,
                   x.name as xa_name,
                   t.hamlet_name as hamlet_display
            FROM members m
            LEFT JOIN organizations t ON m.organization_id = t.id
            LEFT JOIN organizations x ON t.parent_id = x.id
            WHERE 1 = 1
        """
        params = []

        # Org scoping: to_hoi users can see their chi_hoi or members, chi_hoi users can see their org members
        # (role='to_hoi' is used for chi_hoi managers)
        if session.get('role') == 'to_hoi' and session.get('organization_id'):
            query += " AND m.organization_id = %s"
            params.append(session['organization_id'])
        elif session.get('role') == 'chi_hoi' and session.get('organization_id'):
            # chi_hoi can see members from their chi_hoi only
            query += " AND m.organization_id = %s"
            params.append(session['organization_id'])

        if search:
            query += """
                AND (
                    m.full_name LIKE %s
                    OR m.id_number LIKE %s
                    OR m.phone LIKE %s
                    OR m.email LIKE %s
                )
            """
            like_value = f"%{search}%"
            params.extend([like_value, like_value, like_value, like_value])

        if member_type:
            query += " AND m.member_type = %s"
            params.append(member_type)

        if education_level:
            query += " AND m.education_level = %s"
            params.append(education_level)

        if status:
            query += " AND m.status = %s"
            params.append(status)

        if gender:
            query += " AND m.gender = %s"
            params.append(gender)
        
        if org_filter_ids:
            placeholders = ','.join(['%s'] * len(org_filter_ids))
            query += f" AND m.organization_id IN ({placeholders})"
            params.extend(org_filter_ids)

        query += " ORDER BY m.full_name ASC"
        
        # Add pagination: 50 members per page
        offset = (page - 1) * per_page
        
        # Get total count for pagination - reuse the same filter logic
        count_params = []
        count_query = f"SELECT COUNT(*) as total FROM members m WHERE 1=1"
        if session.get('role') == 'to_hoi' and session.get('organization_id'):
            count_query += " AND m.organization_id = %s"
            count_params.append(session['organization_id'])
        elif session.get('role') == 'chi_hoi' and session.get('organization_id'):
            count_query += " AND m.organization_id = %s"
            count_params.append(session['organization_id'])
        # Add filters to count query
        if search:
            count_query += " AND (m.full_name LIKE %s OR m.id_number LIKE %s OR m.phone LIKE %s OR m.email LIKE %s)"
            like_value = f"%{search}%"
            count_params.extend([like_value, like_value, like_value, like_value])
        if member_type:
            count_query += " AND m.member_type = %s"
            count_params.append(member_type)
        if education_level:
            count_query += " AND m.education_level = %s"
            count_params.append(education_level)
        if status:
            count_query += " AND m.status = %s"
            count_params.append(status)
        if gender:
            count_query += " AND m.gender = %s"
            count_params.append(gender)
        if org_filter_ids:
            placeholders = ','.join(['%s'] * len(org_filter_ids))
            count_query += f" AND m.organization_id IN ({placeholders})"
            count_params.extend(org_filter_ids)
        
        cursor.execute(count_query, tuple(count_params))
        total_members = cursor.fetchone()['total']
        total_pages = (total_members + per_page - 1) // per_page

        query += f" LIMIT {per_page} OFFSET {offset}"
        cursor.execute(query, tuple(params))
        members = cursor.fetchall()
        
        # Capitalize all word first letters of hamlet_display (title case)
        for member in members:
            if member.get('hamlet_display'):
                hamlet_name = member['hamlet_display']
                if hamlet_name:
                    member['hamlet_display'] = hamlet_name.title()

        cursor.execute("""
            SELECT id, name as display_name, org_type, parent_id
            FROM organizations
            WHERE org_type IN ('chi_hoi', 'to_hoi')
            ORDER BY org_type DESC, name ASC
        """)
        all_orgs = cursor.fetchall()
        
        # Separate into chi_hoi and to_hoi lists
        member_orgs = [o for o in all_orgs if o['org_type'] == 'chi_hoi']
        to_hoi_orgs = [o for o in all_orgs if o['org_type'] == 'to_hoi']

        cursor.close()
        conn.close()

    return render_template(
        'members.html',
        members=members,
        member_orgs=member_orgs,
        to_hoi_orgs=to_hoi_orgs,
        open_add_member=open_add_member,
        filters={
            'q': search,
            'member_type': member_type,
            'education_level': education_level,
            'status': status,
            'organization_id': organization_id,
            'to_hoi_id': to_hoi_id,
            'organization_id': organization_id,
            'gender': gender
        },
        pagination={
            'page': page,
            'per_page': per_page,
            'total': total_members,
            'total_pages': total_pages
        }
    )

@app.route('/members/add', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'chi_hoi', 'to_hoi')
def add_member():
    """Thêm hội viên"""
    if request.method == 'POST':
        data = request.form.to_dict()
        
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()

            selected_org_id = data.get('organization_id') or None

            # to_hoi users can only add members to their own organization
            # chi_hoi users can add members to their organization
            if session.get('role') == 'to_hoi' and session.get('organization_id'):
                if not selected_org_id or int(selected_org_id) != int(session['organization_id']):
                    cursor.close()
                    conn.close()
                    flash('Bạn chỉ có thể thêm hội viên vào chi hội của mình', 'danger')
                    return redirect(url_for('members_list'))
            elif session.get('role') == 'chi_hoi' and session.get('organization_id'):
                # Check if org_id is chi_hoi
                if not selected_org_id or int(selected_org_id) != int(session['organization_id']):
                    cursor.close()
                    conn.close()
                    flash('Bạn chỉ có thể thêm hội viên vào chi hội của mình', 'danger')
                    return redirect(url_for('members_list'))
                    flash('Bạn chỉ được thêm hội viên vào tổ của mình', 'danger')
                    return redirect(url_for('members_list'))
            
            if selected_org_id:
                cursor.execute("SELECT id, org_type FROM organizations WHERE id = %s", (selected_org_id,))
                selected_org = cursor.fetchone()
                if not selected_org or selected_org['org_type'] != 'to_hoi':
                    cursor.close()
                    conn.close()
                    flash('Chỉ được gán hội viên vào Tổ hội', 'danger')
                    return redirect(url_for('members_list'))
            
            # Hash password nếu có
            password_hash = generate_password_hash(data.get('password', '123456'))
            
            cursor.execute("""
                INSERT INTO members (
                    full_name, date_of_birth, gender, id_number, phone, email, address,
                    education_level, ethnicity, religion, member_type,
                    organization_id, join_date, party_join_date, specialty, politics, password
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                data.get('full_name'),
                data.get('date_of_birth') or None,
                data.get('gender') or 'khac',
                data.get('id_number'),
                data.get('phone'),
                data.get('email'),
                data.get('address'),
                data.get('education_level'),
                data.get('ethnicity'),
                data.get('religion'),
                data.get('member_type') or 'thuong',
                selected_org_id,
                data.get('join_date') or None,
                data.get('party_join_date') or None,
                data.get('specialty'),
                data.get('politics'),
                password_hash
            ))
            
            # Get the new member ID
            member_id = cursor.lastrowid
            
            # Add member to their primary organization in many-to-many table
            if selected_org_id and member_id:
                cursor.execute("""
                    INSERT INTO member_organizations (member_id, organization_id, join_date)
                    VALUES (%s, %s, %s)
                """, (member_id, selected_org_id, data.get('join_date') or None))
            
            conn.commit()
            cursor.close()
            conn.close()
            
            flash('Thêm hội viên thành công', 'success')
            return redirect(url_for('members_list'))
    
    # Lấy danh sách tổ chức (to_hoi - các tộ/nhóm cơ sở)
    conn = get_db_connection()
    organizations = []
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT t.id,
                   CONCAT(h.name, ' → ', x.name, ' → ', c.name, ' → ', t.name) as display_name
            FROM organizations t
            LEFT JOIN organizations c ON t.parent_id = c.id
            LEFT JOIN organizations x ON c.parent_id = x.id
            LEFT JOIN organizations h ON x.parent_id = h.id
            WHERE t.org_type = 'to_hoi'
            ORDER BY h.name, x.name, c.name, t.name
        """)
        organizations = cursor.fetchall()
        cursor.close()
        conn.close()
    
    return render_template('add_member.html', organizations=organizations)

@app.route('/api/members/<int:member_id>')
@login_required
def api_get_member(member_id):
    """API endpoint to get member data as JSON"""
    conn = get_db_connection()
    
    if conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM members WHERE id = %s", (member_id,))
        member = cursor.fetchone()
        
        if not member:
            cursor.close()
            conn.close()
            return jsonify({'error': 'Member not found'}), 404
        
        # Permission check - to_hoi users can only access their own members
        if session.get('role') == 'to_hoi':
            cursor.execute("SELECT organization_id FROM users WHERE id = %s", (session['user_id'],))
            user = cursor.fetchone()
            if user and member['organization_id'] != user['organization_id']:
                cursor.close()
                conn.close()
                return jsonify({'error': 'Permission denied'}), 403
        
        # Get organization name (remove "Hội " prefix if present)
        org_name = ''
        if member['organization_id']:
            cursor.execute("SELECT name FROM organizations WHERE id = %s", (member['organization_id'],))
            org = cursor.fetchone()
            if org:
                org_name = org['name']
                if org_name.startswith('Hội '):
                    org_name = org_name[4:]
        
        cursor.close()
        conn.close()
        
        # Format date_of_birth if it exists
        dob_str = None
        if member['date_of_birth']:
            dob_str = member['date_of_birth'].strftime('%Y-%m-%d')
        
        jd_str = None
        if member['join_date']:
            jd_str = member['join_date'].strftime('%Y-%m-%d')
        
        pjd_str = None
        if member.get('party_join_date'):
            pjd_str = member['party_join_date'].strftime('%Y-%m-%d')
        
        # Format the member data to return
        return jsonify({
            'member': {
                'id': member['id'],
                'full_name': member['full_name'],
                'date_of_birth': dob_str,
                'gender': member['gender'],
                'id_number': member['id_number'],
                'email': member['email'],
                'phone': member['phone'],
                'address': member['address'],
                'education_level': member['education_level'],
                'join_date': jd_str,
                'hamlet_name': org_name,
                'member_type': member['member_type'],
                'status': member['status'],
                'party_join_date': pjd_str,
                'specialty': member.get('specialty'),
                'politics': member.get('politics')
            }
        })
    
    return jsonify({'error': 'Database connection failed'}), 500


@app.route('/api/members/<int:member_id>/organizations', methods=['GET'])
@login_required
def api_member_organizations(member_id):
    """Lấy danh sách tổ chức mà hội viên thuộc về (many-to-many)"""
    orgs = get_member_organizations(member_id)
    
    return jsonify({
        'organizations': orgs
    })


@app.route('/api/members/<int:member_id>/organizations/<int:org_id>', methods=['POST', 'DELETE'])
@login_required
def api_manage_member_organization(member_id, org_id):
    """Thêm hoặc xóa hội viên khỏi một tổ chức"""
    conn = get_db_connection()
    
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    
    cursor = conn.cursor()
    
    # Permission check - chỉ admin hoặc người quản lý tổ chức có thể sửa
    if session.get('role') != 'admin':
        cursor.execute("SELECT organization_id FROM users WHERE id = %s", (session['user_id'],))
        user = cursor.fetchone()
        if not user or user['organization_id'] not in (org_id, None):
            # Also check if user's org is parent of target org
            cursor.execute("""
                SELECT id FROM organizations 
                WHERE id = %s AND (
                    parent_id = %s OR 
                    parent_id IN (SELECT id FROM organizations WHERE parent_id = %s) OR
                    parent_id IN (SELECT id FROM organizations WHERE parent_id IN (SELECT id FROM organizations WHERE parent_id = %s))
                )
            """, (org_id, user['organization_id'], user['organization_id'], user['organization_id']))
            
            if not cursor.fetchone():
                cursor.close()
                conn.close()
                return jsonify({'error': 'Permission denied'}), 403
    
    if request.method == 'POST':
        # Thêm hội viên vào tổ chức
        join_date = request.json.get('join_date') if request.is_json else None
        role = request.json.get('role') if request.is_json else None
        
        if add_member_to_organization(member_id, org_id, join_date, role, conn):
            cursor.close()
            conn.close()
            return jsonify({'success': True, 'message': 'Đã thêm hội viên vào tổ chức'})
        else:
            cursor.close()
            conn.close()
            return jsonify({'error': 'Failed to add member to organization'}), 500
    
    elif request.method == 'DELETE':
        # Xóa hội viên khỏi tổ chức
        if remove_member_from_organization(member_id, org_id, conn):
            cursor.close()
            conn.close()
            return jsonify({'success': True, 'message': 'Đã xóa hội viên khỏi tổ chức'})
        else:
            cursor.close()
            conn.close()
            return jsonify({'error': 'Failed to remove member from organization'}), 500


@app.route('/members/<int:member_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_member(member_id):
    """Sửa thông tin hộ viên"""
    conn = get_db_connection()
    member = None
    organizations = []
    
    if conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM members WHERE id = %s", (member_id,))
        member = cursor.fetchone()
        
        if not member:
            cursor.close()
            conn.close()
            flash('Hộ viên không tồn tại', 'danger')
            return redirect(url_for('members_list'))
        
        # Check permission - to_hoi users can only edit their own members, chi_hoi can edit from their org and child to_hoi
        user_role = session.get('role')
        if user_role == 'to_hoi':
            cursor.execute("SELECT organization_id FROM users WHERE id = %s", (session['user_id'],))
            user = cursor.fetchone()
            if user and member['organization_id'] != user['organization_id']:
                cursor.close()
                conn.close()
                flash('Bạn không có quyền sửa hộ viên này', 'danger')
                return redirect(url_for('members_list'))
        elif user_role == 'chi_hoi':
            cursor.execute("SELECT organization_id FROM users WHERE id = %s", (session['user_id'],))
            user = cursor.fetchone()
            if user:
                # Check if member's org is chi_hoi's org or child to_hoi
                cursor.execute("""
                    SELECT id FROM organizations 
                    WHERE id = %s 
                    AND (
                        id = %s 
                        OR (parent_id = %s AND org_type = 'to_hoi')
                    )
                """, (member['organization_id'], user['organization_id'], user['organization_id']))
                if not cursor.fetchone():
                    cursor.close()
                    conn.close()
                    flash('Bạn không có quyền sửa hộ viên này', 'danger')
                    return redirect(url_for('members_list'))
        
        if request.method == 'POST':
            data = request.form.to_dict()
            
            cursor.execute("""
                UPDATE members 
                SET full_name = %s,
                    date_of_birth = %s,
                    gender = %s,
                    email = %s,
                    phone = %s,
                    id_number = %s,
                    education_level = %s,
                    address = %s,
                    join_date = %s,
                    member_type = %s,
                    status = %s,
                    party_join_date = %s,
                    specialty = %s,
                    politics = %s
                WHERE id = %s
            """, (
                data.get('full_name'),
                data.get('date_of_birth'),
                data.get('gender'),
                data.get('email'),
                data.get('phone'),
                data.get('id_number'),
                data.get('education_level'),
                data.get('address'),
                data.get('join_date'),
                data.get('member_type'),
                data.get('status'),
                data.get('party_join_date'),
                data.get('specialty'),
                data.get('politics'),
                member_id
            ))
            conn.commit()
            cursor.close()
            conn.close()
            
            flash('Cập nhật hộ viên thành công', 'success')
            return redirect(url_for('members_list'))
        
        # Get organizations for dropdown
        if user_role == 'admin':
            cursor.execute("""
                SELECT t.id, t.name, x.name as xa_name
                FROM organizations t
                LEFT JOIN organizations x ON t.parent_id = x.id
                WHERE t.org_type = 'chi_hoi'
                ORDER BY x.name, t.name
            """)
        else:
            cursor.execute("""
                SELECT organization_id FROM users WHERE id = %s
            """, (session['user_id'],))
            user = cursor.fetchone()
            cursor.execute("""
                SELECT t.id, t.name, x.name as xa_name
                FROM organizations t
                LEFT JOIN organizations x ON t.parent_id = x.id
                WHERE t.org_type = 'chi_hoi' AND t.id = %s
                ORDER BY x.name, t.name
            """, (user['organization_id'],))
        
        organizations = cursor.fetchall()
        cursor.close()
        conn.close()
    
    return render_template('edit_member.html', member=member, organizations=organizations)

@app.route('/members/<int:member_id>/delete', methods=['POST'])
@login_required
def delete_member(member_id):
    """Xóa hộ viên"""
    conn = get_db_connection()
    
    if conn:
        cursor = conn.cursor()
        
        # Get member info
        cursor.execute("SELECT * FROM members WHERE id = %s", (member_id,))
        member = cursor.fetchone()
        
        # Check permission - to_hoi users can only delete their own members, chi_hoi can delete from their org and child to_hoi
        user_role = session.get('role')
        if user_role == 'to_hoi':
            cursor.execute("SELECT organization_id FROM users WHERE id = %s", (session['user_id'],))
            user = cursor.fetchone()
            if user and member['organization_id'] != user['organization_id']:
                cursor.close()
                conn.close()
                flash('Bạn không có quyền xóa hộ viên này', 'danger')
                return redirect(url_for('members_list'))
        elif user_role == 'chi_hoi':
            cursor.execute("SELECT organization_id FROM users WHERE id = %s", (session['user_id'],))
            user = cursor.fetchone()
            if user:
                # Check if member's org is chi_hoi's org or child to_hoi
                cursor.execute("""
                    SELECT id FROM organizations 
                    WHERE id = %s 
                    AND (
                        id = %s 
                        OR (parent_id = %s AND org_type = 'to_hoi')
                    )
                """, (member['organization_id'], user['organization_id'], user['organization_id']))
                if not cursor.fetchone():
                    cursor.close()
                    conn.close()
                    flash('Bạn không có quyền xóa hộ viên này', 'danger')
                    return redirect(url_for('members_list'))
        
        # Delete the member
        cursor.execute("DELETE FROM members WHERE id = %s", (member_id,))
        conn.commit()
        cursor.close()
        conn.close()
        
        flash('Xóa hộ viên thành công', 'success')
    
    return redirect(url_for('members_list'))

# Organizations management
@app.route('/organizations')
@login_required
def organizations():
    """Danh sách tổ chức"""
    conn = get_db_connection()
    orgs = []
    hamlet_stats = []
    chi_hoi_summaries = []
    parents = []
    open_add_org = (request.args.get('open_add') or '').strip() == '1'
    search_term = (request.args.get('q') or '').strip()
    selected_xa_id = (request.args.get('xa_id') or '').strip()
    status_filter = (request.args.get('status') or '').strip()
    status_filter = (request.args.get('status') or '').strip()
    xa_tree = []
    chi_hoi_rows = []
    
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
                        SELECT o.*, p.name as parent_name,
                                     CASE
                                         WHEN o.org_type = 'xa' THEN o.name
                                         WHEN o.org_type = 'chi_hoi' THEN p.name
                                         WHEN o.org_type = 'to_hoi' THEN gp.name
                                         ELSE NULL
                                     END as xa_name,
                                     CASE
                                         WHEN o.org_type = 'chi_hoi' THEN o.name
                                         WHEN o.org_type = 'to_hoi' THEN p.name
                                         ELSE NULL
                                     END as chi_hoi_name
            FROM organizations o
            LEFT JOIN organizations p ON o.parent_id = p.id
            LEFT JOIN organizations gp ON p.parent_id = gp.id
            ORDER BY FIELD(o.org_type, 'xa', 'chi_hoi', 'to_hoi'), o.name
        """)
        orgs = cursor.fetchall()

        # Query for to_hoi stats - handles both direct xa→to_hoi and xa→chi_hoi→to_hoi
        cursor.execute("""
            SELECT 
                   COALESCE(x.name, p.name) as xa_name,
                   COALESCE(c.name, '') as chi_hoi_name,
                   t.hamlet_name,
                   COUNT(DISTINCT t.id) as to_hoi_count,
                   COUNT(DISTINCT m.id) as member_count
            FROM organizations t
            LEFT JOIN organizations p ON t.parent_id = p.id
            LEFT JOIN organizations c ON t.parent_id = c.id AND c.org_type = 'chi_hoi'
            LEFT JOIN organizations x ON CASE WHEN c.id IS NOT NULL THEN c.parent_id ELSE t.parent_id END = x.id
            LEFT JOIN members m ON m.organization_id = t.id
            WHERE t.org_type = 'to_hoi' AND t.hamlet_name IS NOT NULL
            GROUP BY xa_name, chi_hoi_name, t.hamlet_name
            ORDER BY xa_name, chi_hoi_name, t.hamlet_name
        """)
        hamlet_stats = cursor.fetchall()

        cursor.execute("""
            SELECT id, name, org_type
            FROM organizations
            ORDER BY FIELD(org_type, 'xa', 'chi_hoi', 'to_hoi'), name
        """)
        parents = cursor.fetchall()

        summary_map = {}
        for row in hamlet_stats:
            key = f"{row['xa_name']}|{row['chi_hoi_name']}"
            if key not in summary_map:
                summary_map[key] = {
                    'xa_name': row['xa_name'],
                    'chi_hoi_name': row['chi_hoi_name'],
                    'hamlet_count': 0,
                    'member_count': 0,
                    'to_hoi_count': 0
                }
            summary_map[key]['hamlet_count'] += 1
            summary_map[key]['member_count'] += int(row['member_count'] or 0)
            summary_map[key]['to_hoi_count'] += int(row['to_hoi_count'] or 0)

        chi_hoi_summaries = list(summary_map.values())
        chi_hoi_summaries.sort(key=lambda item: (item['xa_name'] or '', item['chi_hoi_name'] or ''))

        org_by_id = {item['id']: item for item in orgs}
        chi_hoi_by_xa = {}
        to_hoi_by_xa = {}
        to_hoi_by_chi_hoi = {}
        
        for item in orgs:
            if item['org_type'] == 'chi_hoi':
                chi_hoi_by_xa.setdefault(item['parent_id'], []).append(item)
            elif item['org_type'] == 'to_hoi':
                # Add to both xa and chi_hoi maps
                to_hoi_by_xa.setdefault(item['parent_id'], []).append(item)
                if item['parent_id']:
                    to_hoi_by_chi_hoi.setdefault(item['parent_id'], []).append(item)

        for xa in [item for item in orgs if item['org_type'] == 'xa']:
            children = sorted(chi_hoi_by_xa.get(xa['id'], []), key=lambda row: row['name'] or '')
            xa_tree.append({
                'id': xa['id'],
                'name': xa['name'],
                'children': children
            })

        xa_tree.sort(key=lambda row: row['name'] or '')

        normalized_search = search_term.lower()
        
        # Build hierarchy: Display both chi_hoi and to_hoi (if chi_hoi doesn't exist)
        all_chi_hoi = [row for row in orgs if row['org_type'] == 'chi_hoi']
        all_to_hoi = [row for row in orgs if row['org_type'] == 'to_hoi']
        all_xa = [row for row in orgs if row['org_type'] == 'xa']
        
        # If there's no chi_hoi but there's to_hoi, show to_hoi directly
        show_to_hoi_directly = len(all_chi_hoi) == 0 and len(all_to_hoi) > 0
        
        # Show chi_hoi (or to_hoi if no chi_hoi exists) under xa
        for xa in sorted(all_xa, key=lambda row: row['name'] or ''):
            # Check if xa matches filters
            if selected_xa_id and str(xa['id']) != selected_xa_id:
                continue
            
            # Get children: chi_hoi first, fall back to to_hoi if no chi_hoi
            if show_to_hoi_directly:
                children = [r for r in all_to_hoi if r['parent_id'] == xa['id']]
                child_level = 'Tổ Hội'
            else:
                children = [r for r in all_chi_hoi if r['parent_id'] == xa['id']]
                child_level = 'Chi Hội'
            
            for child in sorted(children, key=lambda row: row['name'] or ''):
                if normalized_search and normalized_search not in (child['name'] or '').lower():
                    continue
                    
                xa_name = xa['name']
                # Capitalize all word first letters of xa_name (title case)
                if xa_name:
                    xa_name = xa_name.title()
                chi_hoi_rows.append({
                    'id': child['id'],
                    'name': child['name'],
                    'level_label': child_level,
                    'xa_name': xa_name,
                    'status': child.get('status', 'active'),
                })

        cursor.close()
        conn.close()
        
        # Apply status filter if specified
        if status_filter:
            chi_hoi_rows = [r for r in chi_hoi_rows if r.get('status', 'active') == status_filter]
    
    return render_template(
        'organizations.html',
        organizations=orgs,
        hamlet_stats=hamlet_stats,
        chi_hoi_summaries=chi_hoi_summaries,
        xa_tree=xa_tree,
        chi_hoi_rows=chi_hoi_rows,
        search_term=search_term,
        selected_xa_id=selected_xa_id,
        status_filter=status_filter,
        parents=parents,
        open_add_org=open_add_org
    )

@app.route('/organization/<int:org_id>')
@login_required
def organization_detail(org_id):
    """Chi tiết tổ hội và danh sách hội viên"""
    conn = get_db_connection()
    organization = None
    members = []
    child_organizations = []
    
    if conn:
        cursor = conn.cursor()
        
        # Get organization info
        cursor.execute("""
            SELECT o.*, p.name as parent_name 
            FROM organizations o
            LEFT JOIN organizations p ON o.parent_id = p.id
            WHERE o.id = %s
        """, (org_id,))
        organization = cursor.fetchone()
        
        if organization:
            # Capitalize all word first letters of hamlet_name (title case)
            if organization.get('hamlet_name'):
                hamlet_name = organization['hamlet_name']
                if hamlet_name:
                    organization['hamlet_name'] = hamlet_name.title()
            
            # Get child organizations (to_hoi for chi_hoi)
            if organization['org_type'] == 'chi_hoi':
                cursor.execute("""
                    SELECT id, name, hamlet_name, address, phone, email, leader_name
                    FROM organizations
                    WHERE parent_id = %s AND org_type = 'to_hoi'
                    ORDER BY name ASC
                """, (org_id,))
                child_organizations = cursor.fetchall()
            
            # Get members: if chi_hoi, get from child to_hoi; if to_hoi, get direct members
            if organization['org_type'] == 'chi_hoi':
                # For chi_hoi, get members from all child to_hoi
                cursor.execute("""
                    SELECT m.id, m.full_name, m.join_date, m.member_type,
                           o.name as organization_name
                    FROM members m
                    LEFT JOIN organizations o ON m.organization_id = o.id
                    WHERE m.organization_id IN (
                        SELECT id FROM organizations WHERE parent_id = %s AND org_type = 'to_hoi'
                    )
                    ORDER BY m.full_name ASC
                """, (org_id,))
            else:
                # For other types (xa, to_hoi), get direct members
                cursor.execute("""
                    SELECT m.id, m.full_name, m.join_date, m.member_type,
                           o.name as organization_name
                    FROM members m
                    LEFT JOIN organizations o ON m.organization_id = o.id
                    WHERE m.organization_id = %s
                    ORDER BY m.full_name ASC
                """, (org_id,))
            
            members = cursor.fetchall()
        
        cursor.close()
        conn.close()
    
    if not organization:
        flash('Tổ chức không tồn tại', 'danger')
        return redirect(url_for('organizations'))
    
    return render_template(
        'organization_detail.html',
        organization=organization,
        members=members,
        child_organizations=child_organizations
    )

@app.route('/organizations/add', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'chi_hoi')
def add_organization():
    """Thêm tổ chức"""
    if request.method == 'POST':
        data = request.form.to_dict()

        org_type = data.get('org_type') or 'xa'
        parent_id = data.get('parent_id') or None
        
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()

            parent = None
            if parent_id:
                try:
                    parent_id = int(parent_id)
                    cursor.execute("SELECT id, org_type FROM organizations WHERE id = %s", (parent_id,))
                    parent = cursor.fetchone()
                except:
                    parent_id = None

            if not validate_org_parent(org_type, parent):
                cursor.close()
                conn.close()
                error_msg = 'Cấu trúc cấp tổ chức không hợp lệ (huyện → xã → chi hội → tổ hội)'
                flash(error_msg, 'danger')
                return redirect(url_for('organizations'))

            cursor.execute("""
                INSERT INTO organizations (name, org_type, parent_id, hamlet_name, address, phone, email, leader_name)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                data.get('name'),
                org_type,
                parent_id,
                data.get('hamlet_name'),
                data.get('address'),
                data.get('phone'),
                data.get('email'),
                data.get('leader_name')
            ))
            conn.commit()
            cursor.close()
            conn.close()
            
            flash('Thêm tổ chức thành công', 'success')
            return redirect(url_for('organizations'))
    
    # Redirect GET request to organizations page
    return redirect(url_for('organizations'))

# Fallback route for AJAX POST requests from browser cache (old JS may still POST to /organizations/<id>)
@app.route('/organizations/<int:org_id>', methods=['POST'])
@login_required
@admin_required
def update_organization_ajax(org_id):
    """Fallback handler for AJAX organization updates (redirects to edit_organization)"""
    # Forward to the actual edit_organization handler
    return edit_organization(org_id)

@app.route('/organizations/<int:org_id>/edit', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'chi_hoi')
def edit_organization(org_id):
    """Sửa tổ chức"""
    conn = get_db_connection()
    org = None
    parents = []
    
    if conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM organizations WHERE id = %s", (org_id,))
        org = cursor.fetchone()
        
        if not org:
            cursor.close()
            conn.close()
            if request.is_json or request.accept_mimetypes.accept_json and not request.accept_mimetypes.accept_html:
                return jsonify({'error': 'Tổ chức không tồn tại'}), 404
            flash('Tổ chức không tồn tại', 'danger')
            return redirect(url_for('organizations'))
        
        # Check permission - chi_hoi/to_hoi users can only edit their own organization
        if (session.get('role') in ('to_hoi', 'chi_hoi')) and session.get('organization_id'):
            user_org_id = session['organization_id']
            # Check if editing their own org
            if org_id != user_org_id:
                cursor.close()
                conn.close()
                flash('Bạn không có quyền sửa chi hội này', 'danger')
                return redirect(url_for('organizations'))
        
        if request.method == 'POST':
            data = request.form.to_dict()
            
            cursor.execute("""
                UPDATE organizations 
                SET name = %s,
                    address = %s,
                    phone = %s,
                    email = %s,
                    leader_name = %s
                WHERE id = %s
            """, (
                data.get('name'),
                data.get('address'),
                data.get('phone'),
                data.get('email'),
                data.get('leader_name'),
                org_id
            ))
            conn.commit()
            cursor.close()
            conn.close()
            
            # Check if AJAX request
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': True, 'message': 'Cập nhật tổ chức thành công'})
            
            flash('Cập nhật tổ chức thành công', 'success')
            return redirect(url_for('organizations'))
        
        cursor.execute("""
            SELECT id, name, org_type
            FROM organizations
            ORDER BY FIELD(org_type, 'xa', 'chi_hoi', 'to_hoi'), name
        """)
        parents = cursor.fetchall()
        cursor.close()
        conn.close()
    
    return render_template('edit_organization.html', organization=org, parents=parents)

@app.route('/api/organizations/<int:org_id>', methods=['GET'])
@login_required
def api_get_organization(org_id):
    """API lấy thông tin tổ chức"""
    conn = get_db_connection()
    
    if conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM organizations WHERE id = %s", (org_id,))
        org = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if org:
            return jsonify({
                'id': org['id'],
                'name': org['name'],
                'org_type': org['org_type'],
                'hamlet_name': org['hamlet_name'] or '',
                'address': org['address'] or '',
                'phone': org['phone'] or '',
                'email': org['email'] or '',
                'leader_name': org['leader_name'] or ''
            })
        else:
            return jsonify({'error': 'Tổ chức không tồn tại'}), 404
    
    return jsonify({'error': 'Lỗi kết nối'}), 500

@app.route('/api/organizations/<int:chi_hoi_id>/to_hoi_list', methods=['GET'])
@login_required
def api_get_to_hoi_by_chi_hoi(chi_hoi_id):
    """API lấy danh sách tổ hội theo chi hội"""
    conn = get_db_connection()
    
    if conn:
        cursor = conn.cursor()
        try:
            # Get all tổ hội that have this chi_hoi as parent
            cursor.execute("""
                SELECT id, name as display_name, org_type
                FROM organizations
                WHERE parent_id = %s AND org_type = 'to_hoi'
                ORDER BY name ASC
            """, (chi_hoi_id,))
            to_hoies = cursor.fetchall()
            cursor.close()
            conn.close()
            
            return jsonify({
                'to_hoies': [{'id': t['id'], 'name': t['display_name']} for t in to_hoies]
            })
        except Exception as e:
            cursor.close()
            conn.close()
            return jsonify({'error': str(e)}), 500
    
    return jsonify({'error': 'Lỗi kết nối'}), 500

@app.route('/organizations/<int:org_id>/delete', methods=['POST'])
@login_required
@roles_required('admin', 'chi_hoi')
def delete_organization(org_id):
    """Xóa tổ chức"""
    conn = get_db_connection()
    
    if conn:
        cursor = conn.cursor()
        
        # Check permission - chi_hoi/to_hoi users can only delete their own organization
        if (session.get('role') in ('to_hoi', 'chi_hoi')) and session.get('organization_id'):
            user_org_id = session['organization_id']
            # Check if deleting their own org
            if org_id != user_org_id:
                cursor.close()
                conn.close()
                flash('Bạn không có quyền xóa chi hội này', 'danger')
                return redirect(url_for('organizations'))
        
        # Check if organization has members
        cursor.execute("SELECT COUNT(*) as count FROM members WHERE organization_id = %s", (org_id,))
        member_count = cursor.fetchone()['count']
        
        if member_count > 0:
            cursor.close()
            conn.close()
            flash(f'Không thể xóa - Tổ chức này đang có {member_count} hộ viên', 'danger')
            return redirect(url_for('organizations'))
        
        # Check if organization has children
        cursor.execute("SELECT COUNT(*) as count FROM organizations WHERE parent_id = %s", (org_id,))
        child_count = cursor.fetchone()['count']
        
        if child_count > 0:
            cursor.close()
            conn.close()
            flash(f'Không thể xóa - Tổ chức này đang có {child_count} tổ chức con', 'danger')
            return redirect(url_for('organizations'))
        
        # Delete the organization
        cursor.execute("DELETE FROM organizations WHERE id = %s", (org_id,))
        conn.commit()
        cursor.close()
        conn.close()
        
        flash('Xóa tổ chức thành công', 'success')
    
    return redirect(url_for('organizations'))

@app.route('/organizations/<int:org_id>/stop', methods=['POST'])
@login_required
@roles_required('admin', 'chi_hoi', 'to_hoi')
def stop_organization(org_id):
    """Dừng hoạt động tổ chức"""
    conn = get_db_connection()
    
    if conn:
        cursor = conn.cursor()
        
        try:
            # Check permission - chi_hoi/to_hoi users can only stop their own organization
            if (session.get('role') in ('to_hoi', 'chi_hoi')) and session.get('organization_id'):
                user_org_id = session['organization_id']
                if org_id != user_org_id:
                    cursor.close()
                    conn.close()
                    flash('Bạn không có quyền dừng hoạt động chi hội này', 'danger')
                    return redirect(url_for('organizations'))
            
            # Try to update status column
            try:
                cursor.execute("UPDATE organizations SET status = 'inactive' WHERE id = %s", (org_id,))
                conn.commit()
            except:
                # If status column doesn't exist, add it first
                cursor.execute("ALTER TABLE organizations ADD COLUMN status VARCHAR(50) DEFAULT 'active'")
                conn.commit()
                cursor.execute("UPDATE organizations SET status = 'inactive' WHERE id = %s", (org_id,))
                conn.commit()
            
            flash('Đã dừng hoạt động tổ chức', 'success')
        except Exception as e:
            print(f"Error stopping organization: {e}")
            flash(f'Lỗi: {str(e)}', 'danger')
        finally:
            cursor.close()
            conn.close()
    
    return redirect(url_for('organizations'))

@app.route('/organizations/<int:org_id>/reactivate', methods=['POST'])
@login_required
@roles_required('admin', 'chi_hoi', 'to_hoi')
def reactivate_organization(org_id):
    """Hoạt động lại tổ chức"""
    conn = get_db_connection()
    
    if conn:
        cursor = conn.cursor()
        
        try:
            # Check permission - chi_hoi/to_hoi users can only reactivate their own organization
            if (session.get('role') in ('to_hoi', 'chi_hoi')) and session.get('organization_id'):
                user_org_id = session['organization_id']
                if org_id != user_org_id:
                    cursor.close()
                    conn.close()
                    flash('Bạn không có quyền hoạt động lại chi hội này', 'danger')
                    return redirect(url_for('organizations'))
            
            # Update status to active
            cursor.execute("UPDATE organizations SET status = 'active' WHERE id = %s", (org_id,))
            conn.commit()
            
            flash('Đã hoạt động lại tổ chức', 'success')
        except Exception as e:
            print(f"Error reactivating organization: {e}")
            flash(f'Lỗi: {str(e)}', 'danger')
        finally:
            cursor.close()
            conn.close()
    
    return redirect(url_for('organizations'))

# Reports
@app.route('/reports/member-summary')
@login_required
def reports_member_summary():
    """Thống kê số lượng hội viên"""
    conn = get_db_connection()
    summary = {
        'total_members': 0,
        'active_members': 0,
        'inactive_members': 0,
        'dang_vien_members': 0,
        'nong_cot_members': 0
    }
    by_xa = []

    if conn:
        cursor = conn.cursor()

        cursor.execute("SELECT COUNT(*) as count FROM members")
        summary['total_members'] = cursor.fetchone()['count']

        cursor.execute("SELECT COUNT(*) as count FROM members WHERE status = 'active'")
        summary['active_members'] = cursor.fetchone()['count']

        cursor.execute("SELECT COUNT(*) as count FROM members WHERE status != 'active'")
        summary['inactive_members'] = cursor.fetchone()['count']

        cursor.execute("SELECT COUNT(*) as count FROM members WHERE member_type = 'dang_vien'")
        summary['dang_vien_members'] = cursor.fetchone()['count']

        cursor.execute("SELECT COUNT(*) as count FROM members WHERE member_type = 'nong_cot'")
        summary['nong_cot_members'] = cursor.fetchone()['count']

        cursor.execute("""
            SELECT chi.id, chi.name, COUNT(m.id) as member_count
            FROM members m
            LEFT JOIN organizations to_hoi ON m.organization_id = to_hoi.id
            LEFT JOIN organizations chi ON to_hoi.parent_id = chi.id
            WHERE chi.org_type = 'chi_hoi'
            GROUP BY chi.id, chi.name
            ORDER BY member_count DESC, chi.name
        """)
        by_xa = cursor.fetchall()

        cursor.close()
        conn.close()

    return render_template(
        'reports_member_summary.html',
        summary=summary,
        by_xa=by_xa
    )


@app.route('/reports/member-detail')
@login_required
def reports_member_detail():
    """Chi tiết tổ chức"""
    q = (request.args.get('q') or '').strip()
    org_type = (request.args.get('org_type') or '').strip()

    conn = get_db_connection()
    organizations = []
    export_organizations = []

    if conn:
        cursor = conn.cursor()
        query = """
            SELECT o.id,
                   o.name,
                   o.org_type,
                   o.hamlet_name,
                   p.name as parent_name,
                   COUNT(DISTINCT child.id) as child_count,
                   COUNT(DISTINCT m.id) as member_count
            FROM organizations o
            LEFT JOIN organizations p ON o.parent_id = p.id
            LEFT JOIN organizations child ON child.parent_id = o.id
            LEFT JOIN members m ON (
                m.organization_id = o.id 
                OR m.organization_id IN (
                    SELECT id FROM organizations 
                    WHERE parent_id = o.id
                    UNION
                    SELECT child_orgs.id FROM organizations child_orgs
                    LEFT JOIN organizations parent_orgs ON child_orgs.parent_id = parent_orgs.id
                    WHERE parent_orgs.parent_id = o.id
                )
            )
            WHERE 1 = 1
        """
        params = []

        if q:
            query += """
                AND (
                    o.name LIKE %s
                    OR p.name LIKE %s
                )
            """
            search_like = f"%{q}%"
            params.extend([search_like, search_like])

        if org_type:
            query += " AND o.org_type = %s"
            params.append(org_type)

        query += """
            GROUP BY o.id, o.name, o.org_type, o.hamlet_name, p.name
            ORDER BY FIELD(o.org_type, 'xa', 'chi_hoi', 'to_hoi'), o.name
        """

        cursor.execute(query, tuple(params))
        organizations = cursor.fetchall()
        
        # Capitalize all word first letters of hamlet_name (title case)
        for org in organizations:
            if org.get('hamlet_name'):
                hamlet_name = org['hamlet_name']
                if hamlet_name:
                    org['hamlet_name'] = hamlet_name.title()
        
        # Get all organizations for export dropdown (Chi Hội and Tổ Hội only)
        cursor.execute("""
            SELECT id, name FROM organizations 
            WHERE org_type IN ('chi_hoi', 'to_hoi')
            ORDER BY name
        """)
        export_organizations = cursor.fetchall()
        
        cursor.close()
        conn.close()

    return render_template(
        'reports_member_detail.html',
        organizations=organizations,
        export_organizations=export_organizations,
        filters={
            'q': q,
            'org_type': org_type
        }
    )


# Import from Excel
@app.route('/members/import', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'chi_hoi', 'to_hoi')
def import_members():
    """Nhập hội viên từ file Excel"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Vui lòng chọn file', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('Vui lòng chọn file', 'danger')
            return redirect(request.url)
        
        if not allowed_file(file.filename):
            flash('Chỉ chấp nhận file Excel (.xlsx, .xls)', 'danger')
            return redirect(request.url)
        
        try:
            # Save file temporarily
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            # Load workbook
            wb = load_workbook(filepath)
            ws = wb.active
            
            conn = get_db_connection()
            cursor = conn.cursor()
            
            success_count = 0
            error_messages = []
            
            # Skip header row
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Map columns: full_name, date_of_birth, gender, id_number, phone, email, address, education_level, ethnicity, religion, member_type, organization_id
                    full_name = row[0]
                    date_of_birth = row[1]
                    gender = row[2] or 'khac'
                    id_number = row[3]
                    phone = row[4]
                    email = row[5]
                    address = row[6]
                    education_level = row[7]
                    ethnicity = row[8]
                    religion = row[9]
                    member_type = row[10] or 'thuong'
                    organization_id = row[11]
                    
                    if not full_name or not organization_id:
                        error_messages.append(f"Dòng {row_idx}: Tên hội viên và Tổ chức không được để trống")
                        continue
                    
                    # Check organization exists and is 'chi_hoi'
                    cursor.execute("SELECT id, org_type, hamlet_name FROM organizations WHERE id = %s", (organization_id,))
                    org = cursor.fetchone()
                    if not org:
                        error_messages.append(f"Dòng {row_idx}: Tổ chức không tồn tại")
                        continue
                    if org['org_type'] != 'chi_hoi':
                        error_messages.append(f"Dòng {row_idx}: Chỉ được gán hội viên vào Chi hội")
                        continue
                    
                    # to_hoi/chi_hoi users can only add to their own organization
                    if session.get('role') in ('to_hoi', 'chi_hoi') and session.get('organization_id'):
                        if int(organization_id) != int(session['organization_id']):
                            error_messages.append(f"Dòng {row_idx}: Bạn chỉ được nhập hội viên vào chi hội của mình")
                            continue
                    
                    # Parse date
                    parsed_date = None
                    if date_of_birth:
                        try:
                            if isinstance(date_of_birth, datetime):
                                parsed_date = date_of_birth.strftime('%Y-%m-%d')
                            else:
                                parsed_date = str(date_of_birth)
                        except:
                            pass
                    
                    password_hash = generate_password_hash('123456')
                    
                    cursor.execute("""
                        INSERT INTO members (
                            full_name, date_of_birth, gender, id_number, phone, email, address,
                            education_level, ethnicity, religion, member_type,
                            organization_id, join_date, password, status
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        full_name, parsed_date, gender, id_number, phone, email, address,
                        education_level, ethnicity, religion, member_type,
                        organization_id, datetime.now().strftime('%Y-%m-%d'), password_hash, 'active'
                    ))
                    success_count += 1
                
                except Exception as e:
                    error_messages.append(f"Dòng {row_idx}: Lỗi - {str(e)}")
            
            conn.commit()
            cursor.close()
            conn.close()
            
            # Delete uploaded file
            os.remove(filepath)
            
            flash(f'Nhập {success_count} hội viên thành công', 'success')
            if error_messages:
                flash('Có lỗi: ' + '; '.join(error_messages[:5]), 'warning')
            
            return redirect(url_for('members_list'))
        
        except Exception as e:
            flash(f'Lỗi khi xử lý file: {str(e)}', 'danger')
            return redirect(request.url)
    
    return render_template('import_members.html')


@app.route('/members/import-batch', methods=['GET', 'POST'])
@login_required
@roles_required('admin')
def import_members_batch():
    """Nhập hội viên từ tất cả file Excel trong thư mục"""
    if request.method == 'GET':
        # GET: Chỉ render template, không cần query database
        return render_template('import_batch.html')
    
    if request.method == 'POST':        
        excel_folder = 'exel hội viên theo từng ấp'
        
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            total_success = 0
            total_errors = []
            
            # Lấy danh sách tất cả file Excel trong thư mục
            excel_files = sorted(glob.glob(f'{excel_folder}/*.xlsx') + glob.glob(f'{excel_folder}/*.xls'))
            
            if not excel_files:
                flash('Không tìm thấy file Excel nào trong thư mục', 'danger')
                return redirect(url_for('members_list'))
            
            for filepath in excel_files:
                try:
                    # Extract hamlet name from filename
                    # "hội ấp bà trầm.xlsx" → "bà trầm"
                    filename = os.path.basename(filepath)
                    hamlet_name = filename.replace('hội ấp ', '').replace('.xlsx', '').replace('.xls', '').strip()
                    
                    # Load workbook
                    wb = load_workbook(filepath)
                    ws = wb.active
                    
                    file_success = 0
                    file_errors = []
                    
                    # Skip header row, start from row 2
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        try:
                            # Column mapping for the specific Excel structure:
                            # Col 1: STT (skip)
                            # Col 2: HỌ VÀ TÊN (full_name)
                            # Col 3: GIỚI TÍNH (gender)
                            # Col 4: NĂM SINH (year)
                            # Col 5: EMAIL (email)
                            # Col 6: CHỨC DANH (position/member_type)
                            # Col 7: NGÀY VÀO HỘI (join_date)
                            # Col 8: NGÀNH NGHỀ (profession)
                            
                            stt = row[0]
                            full_name = row[1] if len(row) > 1 else None
                            gender = row[2] if len(row) > 2 else 'khac'
                            year = row[3] if len(row) > 3 else None
                            email = row[4] if len(row) > 4 else None
                            chuc_danh = row[5] if len(row) > 5 else None
                            join_date = row[6] if len(row) > 6 else None
                            profession = row[7] if len(row) > 7 else None
                            
                            # Skip empty rows
                            if not full_name:
                                continue
                            
                            # Map CHỨC DANH to member_type
                            member_type = 'thuong'  # default
                            if chuc_danh:
                                chuc_danh_lower = str(chuc_danh).lower()
                                if 'đảng viên' in chuc_danh_lower or 'đảng' in chuc_danh_lower:
                                    member_type = 'dang_vien'
                                elif 'nòng cốt' in chuc_danh_lower or 'nòng' in chuc_danh_lower:
                                    member_type = 'nong_cot'
                            
                            # Convert year to date (YYYY → YYYY-01-01)
                            date_of_birth = None
                            if year:
                                try:
                                    year_str = str(int(float(year)))
                                    date_of_birth = f'{year_str}-01-01'
                                except:
                                    pass
                            
                            # Convert join_date
                            parsed_join_date = None
                            if join_date:
                                try:
                                    if isinstance(join_date, datetime):
                                        parsed_join_date = join_date.strftime('%Y-%m-%d')
                                    else:
                                        parsed_join_date = str(join_date)[:10]
                                except:
                                    parsed_join_date = datetime.now().strftime('%Y-%m-%d')
                            
                            # Find organization by hamlet_name
                            cursor.execute("""
                                SELECT id, hamlet_name FROM organizations 
                                WHERE org_type = 'to_hoi' AND hamlet_name = %s
                                LIMIT 1
                            """, (hamlet_name,))
                            org_result = cursor.fetchone()
                            
                            if not org_result:
                                file_errors.append(f"Dòng {row_idx}: Không tìm thấy tổ hội cho ấp '{hamlet_name}'")
                                continue
                            
                            organization_id = org_result['id']
                            # Get hamlet_name from organization to ensure consistency
                            hamlet_name_from_org = org_result['hamlet_name']
                            
                            # Normalize gender
                            if gender:
                                gender_lower = str(gender).lower()
                                if 'nam' in gender_lower:
                                    gender = 'nam'
                                elif 'nữ' in gender_lower:
                                    gender = 'nu'
                                else:
                                    gender = 'khac'
                            else:
                                gender = 'khac'
                            
                            password_hash = generate_password_hash('123456')
                            
                            # Insert member
                            cursor.execute("""
                                INSERT INTO members (
                                    full_name, date_of_birth, gender, phone, email,
                                    member_type, organization_id, join_date, 
                                    password, status
                                )
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """, (
                                full_name, date_of_birth, gender, None, email,
                                member_type, organization_id, 
                                parsed_join_date or datetime.now().strftime('%Y-%m-%d'), 
                                password_hash, 'active'
                            ))
                            file_success += 1
                        
                        except Exception as e:
                            file_errors.append(f"Dòng {row_idx}: {str(e)[:50]}")
                    
                    if file_errors:
                        total_errors.append(f"File '{hamlet_name}': {len(file_errors)} lỗi")
                    
                    total_success += file_success
                
                except Exception as e:
                    total_errors.append(f"File '{filename}': Lỗi đọc file - {str(e)[:50]}")
            
            conn.commit()
            cursor.close()
            conn.close()
            
            flash(f'✓ Nhập thành công {total_success} hội viên từ {len(excel_files)} file Excel', 'success')
            if total_errors:
                for err in total_errors[:3]:
                    flash(f'⚠ {err}', 'warning')
            
            return redirect(url_for('members_list'))
        
        except Exception as e:
            flash(f'Lỗi: {str(e)}', 'danger')
            return redirect(url_for('members_list'))
    
    return render_template('import_batch.html')


# Error handlers
@app.route('/export/members/all', methods=['GET'])
@login_required
def export_members_all():
    """Export all members to Excel"""
    conn = get_db_connection()
    
    if not conn:
        flash('Lỗi kết nối database', 'danger')
        return redirect(url_for('reports_member_summary'))
    
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT m.*,
                   t.name as chi_hoi_name,
                   x.name as xa_name
            FROM members m
            LEFT JOIN organizations t ON m.organization_id = t.id
            LEFT JOIN organizations x ON t.parent_id = x.id
            ORDER BY m.full_name ASC
        """)
        members = cursor.fetchall()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Hội viên"
        
        # Add headers
        headers = ['Họ tên', 'Ngày sinh', 'Giới tính', 'Số CMND/CCCD', 'Điện thoại', 'Email', 
                   'Dân tộc', 'Tôn giáo', 'Phân loại', 'Xã', 'Chi hội', 
                   'Trình độ', 'Địa chỉ', 'Trạng thái', 'Ngành nghề', 'Chuyên môn', 'Chính trị']
        ws.append(headers)
        
        # Add data rows
        for member in members:
            row = [
                member.get('full_name', ''),
                member.get('date_of_birth', ''),
                'Nam' if member.get('gender') == 'nam' else 'Nữ' if member.get('gender') == 'nu' else 'Khác',
                member.get('id_number', ''),
                member.get('phone', ''),
                member.get('email', ''),
                member.get('ethnicity', ''),
                member.get('religion', ''),
                'Đảng viên' if member.get('member_type') == 'dang_vien' else 'Nòng cốt' if member.get('member_type') == 'nong_cot' else 'Thường',
                member.get('xa_name', ''),
                member.get('chi_hoi_name', ''),
                member.get('education_level', ''),
                member.get('address', ''),
                'Hoạt động' if member.get('status') == 'active' else 'Không hoạt động' if member.get('status') == 'inactive' else 'Bị khóa',
                member.get('occupation', ''),
                member.get('specialty', ''),
                member.get('politics', '')
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        cursor.close()
        conn.close()
        
        # Create a response with the BytesIO object
        return send_file(
            io.BytesIO(stream.getvalue()),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'danh_sach_hoi_vien_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        print(f'Error in export_members_all: {str(e)}')
        print(f'Traceback: ', exc_info=True)
        flash(f'Lỗi xuất Excel: {str(e)}', 'danger')
        return redirect(url_for('reports_member_summary'))

@app.route('/export/members/organization/<int:org_id>', methods=['GET'])
@login_required
def export_members_by_organization(org_id):
    """Export members of a specific organization to Excel"""
    conn = get_db_connection()
    
    if not conn:
        flash('Lỗi kết nối database', 'danger')
        return redirect(url_for('reports_member_summary'))
    
    try:
        cursor = conn.cursor()
        
        # Get organization info
        cursor.execute("SELECT name, org_type FROM organizations WHERE id = %s", (org_id,))
        org = cursor.fetchone()
        
        if not org:
            flash('Tổ chức không tồn tại', 'danger')
            return redirect(url_for('reports_member_summary'))
        
        # Get members
        cursor.execute("""
            SELECT m.*,
                   t.name as chi_hoi_name,
                   x.name as xa_name
            FROM members m
            LEFT JOIN organizations t ON m.organization_id = t.id
            LEFT JOIN organizations x ON t.parent_id = x.id
            WHERE m.organization_id = %s
            ORDER BY m.full_name ASC
        """, (org_id,))
        members = cursor.fetchall()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Hội viên"
        
        # Add title
        ws['A1'] = f"Danh sách hội viên - {org['name']}"
        ws['A1'].font = ws['A1'].font.copy()
        
        # Add headers
        headers = ['Họ tên', 'Ngày sinh', 'Giới tính', 'Số CMND/CCCD', 'Điện thoại', 'Email', 
                   'Dân tộc', 'Tôn giáo', 'Phân loại', 'Xã', 'Chi hội', 
                   'Trình độ', 'Địa chỉ', 'Trạng thái', 'Ngành nghề', 'Chuyên môn', 'Chính trị']
        ws.append(headers)
        
        # Add data rows
        for member in members:
            row = [
                member.get('full_name', ''),
                member.get('date_of_birth', ''),
                'Nam' if member.get('gender') == 'nam' else 'Nữ' if member.get('gender') == 'nu' else 'Khác',
                member.get('id_number', ''),
                member.get('phone', ''),
                member.get('email', ''),
                member.get('ethnicity', ''),
                member.get('religion', ''),
                'Đảng viên' if member.get('member_type') == 'dang_vien' else 'Nòng cốt' if member.get('member_type') == 'nong_cot' else 'Thường',
                member.get('xa_name', ''),
                member.get('chi_hoi_name', ''),
                member.get('education_level', ''),
                member.get('address', ''),
                'Hoạt động' if member.get('status') == 'active' else 'Không hoạt động' if member.get('status') == 'inactive' else 'Bị khóa',
                member.get('occupation', ''),
                member.get('specialty', ''),
                member.get('politics', '')
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        cursor.close()
        conn.close()
        
        # Create a response with the BytesIO object
        return send_file(
            io.BytesIO(stream.getvalue()),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'danh_sach_hoi_vien_{org["name"]}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        print(f'Error in export_members_org: {str(e)}')
        print(f'Traceback: ', exc_info=True)
        flash(f'Lỗi xuất Excel: {str(e)}', 'danger')
        return redirect(url_for('reports_member_summary'))

@app.errorhandler(404)
def not_found(error):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('500.html'), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
